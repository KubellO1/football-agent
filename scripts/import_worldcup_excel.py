"""从 Excel 导入 World Cup 2026 数据（比赛/球队/赛事/比分/Bet365 赔率）。

设计约束（严格遵守）：
- **只经既有仓储写入**，绝不直接写 SQL、绝不绕过仓储：
  CompetitionRepository / TeamRepository / FixtureRepository /
  OddsSnapshotRepository（Bet365 需要一个 Bookmaker，故也用 BookmakerRepository
  做 get-or-create——仍是既有仓储，未绕过）。
- **幂等**：重复运行不产生重复行。参考数据按 (external_source, external_id)
  get-or-create；比赛按同键 update-or-create；赔率用 add_if_absent（ON CONFLICT）。
- **external_id 策略**：Excel 若给了 external id 就用；否则由
  competition + season + date + home + away 生成**确定性** id（sha1）。
- external_source 固定为 "worldcup-excel"，与真实 API-Football 数据隔离、便于清理。

关于 xG / Shots / Shots on target：领域层**没有**对应的实体/表/仓储来存放「每场
比赛的统计」（predictions 表里的 xG 是模型产出，不是导入口径）。在「只用既有仓储、
不建新表」的约束下，这些列会被**读取并计数，但不落库**，并在结尾明确报告。要持久化
它们需要新增 MatchStats 实体 + 表 + 仓储 + 迁移（超出本次约束）。

预期表结构（每个 worksheet 为一批「每行一场比赛」的数据；宽松匹配列名，大小写/
空格/下划线不敏感；缺列即视为空）：
    Competition, Season, Date, Home, Away, HomeScore, AwayScore,
    B365H, B365D, B365A, xG_Home, xG_Away, Shots_Home, Shots_Away,
    SoT_Home, SoT_Away
    （Competition 缺失时用 worksheet 名兜底；External_Id 列可选）

用法：
    python scripts/import_worldcup_excel.py WorldCup2026.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import re
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook

from app.config.settings import get_settings
from app.core.logging import get_logger
from app.database.session import Database
from app.models.entities.bookmaker import Bookmaker
from app.models.entities.competition import Competition
from app.models.entities.enums import MatchStatus
from app.models.entities.fixture import Fixture
from app.models.entities.odds_snapshot import OddsSnapshot
from app.models.entities.team import Team
from app.models.value_objects.markets import MarketType, Selection
from app.models.value_objects.odds import Odds
from app.models.value_objects.score import Score
from app.repositories.sqlalchemy.fixture_repository import SqlAlchemyFixtureRepository
from app.repositories.sqlalchemy.odds_snapshot_repository import SqlAlchemyOddsSnapshotRepository
from app.repositories.sqlalchemy.reference_repositories import (
    SqlAlchemyBookmakerRepository,
    SqlAlchemyCompetitionRepository,
    SqlAlchemyTeamRepository,
)

logger = get_logger(__name__)

SOURCE = "worldcup-excel"
BET365_KEY = "bet365"

# 列名别名（均以 normalize_key 归一化：仅保留小写字母数字）。
ALIASES: dict[str, list[str]] = {
    "external_id": ["externalid", "matchid", "fixtureid", "id"],
    "competition": ["competition", "league", "comp", "tournament"],
    "season": ["season", "year"],
    "date": ["date", "kickoff", "datetime", "dateutc", "matchdate"],
    "home": ["home", "hometeam", "hometeamname"],
    "away": ["away", "awayteam", "awayteamname"],
    "home_score": ["homescore", "fthg", "hgft", "hg", "scorehome", "homegoals", "goalshome"],
    "away_score": ["awayscore", "ftag", "agft", "ag", "scoreaway", "awaygoals", "goalsaway"],
    "b365_home": ["b365h", "bet365h", "bet365home", "b365home", "oddshome", "homeodds"],
    "b365_draw": ["b365d", "bet365d", "bet365draw", "b365draw", "oddsdraw", "drawodds"],
    "b365_away": ["b365a", "bet365a", "bet365away", "b365away", "oddsaway", "awayodds"],
    "xg_home": ["xghome", "homexg", "xgh", "hxg"],
    "xg_away": ["xgaway", "awayxg", "xga", "axg"],
    "shots_home": ["shotshome", "homeshots", "hs"],
    "shots_away": ["shotsaway", "awayshots"],
    "sot_home": ["sothome", "homesot", "shotsontargethome", "hst"],
    "sot_away": ["sotaway", "awaysot", "shotsontargetaway", "ast"],
}


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower()) if value is not None else ""


def deterministic_id(*parts: Any) -> str:
    raw = "|".join(normalize_key(p) for p in parts)
    return "wc-" + hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def parse_date(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=UTC)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=UTC)
        except ValueError:
            continue
    return None


def to_int(value: Any) -> int | None:
    try:
        return int(float(value)) if value not in (None, "") else None
    except (ValueError, TypeError):
        return None


def to_decimal(value: Any) -> Decimal | None:
    try:
        return Decimal(str(value)) if value not in (None, "") else None
    except (InvalidOperation, ValueError, TypeError):
        return None


@dataclass
class Counters:
    matches_imported: int = 0
    scores_imported: int = 0
    teams_created: int = 0
    teams_updated: int = 0
    odds_imported: int = 0
    skipped_duplicates: int = 0
    rows_invalid: int = 0
    stats_rows_read: int = 0  # xG/shots/SoT 读取但不落库


def build_column_map(header: tuple[Any, ...]) -> dict[str, int]:
    """把表头映射为 {字段: 列下标}。"""
    normalized = {normalize_key(h): i for i, h in enumerate(header) if h is not None}
    col: dict[str, int] = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                col[field] = normalized[alias]
                break
    return col


class WorldCupExcelImporter:
    """经既有仓储、幂等地导入一份 World Cup Excel。"""

    def __init__(self, session) -> None:  # type: ignore[no-untyped-def]
        self._competitions = SqlAlchemyCompetitionRepository(session)
        self._teams = SqlAlchemyTeamRepository(session)
        self._fixtures = SqlAlchemyFixtureRepository(session)
        self._bookmakers = SqlAlchemyBookmakerRepository(session)
        self._odds = SqlAlchemyOddsSnapshotRepository(session)
        self._counters = Counters()
        self._comp_cache: dict[str, Competition] = {}
        self._team_cache: dict[str, Team] = {}
        self._bet365: Bookmaker | None = None

    async def _bet365_bookmaker(self) -> Bookmaker:
        if self._bet365 is None:
            existing = await self._bookmakers.get_by_external_id(SOURCE, BET365_KEY)
            self._bet365 = existing or await self._bookmakers.add(
                Bookmaker(name="Bet365", external_id=BET365_KEY, external_source=SOURCE)
            )
        return self._bet365

    async def _get_or_create_competition(self, name: str, country: str | None) -> Competition:
        ext = deterministic_id("competition", name)
        if ext in self._comp_cache:
            return self._comp_cache[ext]
        existing = await self._competitions.get_by_external_id(SOURCE, ext)
        comp = existing or await self._competitions.add(
            Competition(name=name, country=country or "", external_id=ext, external_source=SOURCE)
        )
        self._comp_cache[ext] = comp
        return comp

    async def _upsert_team(self, name: str, country: str | None) -> Team:
        ext = deterministic_id("team", name)
        if ext in self._team_cache:
            return self._team_cache[ext]
        existing = await self._teams.get_by_external_id(SOURCE, ext)
        if existing is not None:
            existing.name = name
            existing.country = country or existing.country
            team = await self._teams.update(existing)
            self._counters.teams_updated += 1
        else:
            team = await self._teams.add(
                Team(name=name, country=country, external_id=ext, external_source=SOURCE)
            )
            self._counters.teams_created += 1
        self._team_cache[ext] = team
        return team

    async def _upsert_fixture(
        self,
        ext: str,
        competition_id: UUID,
        home_id: UUID,
        away_id: UUID,
        kickoff: datetime,
        score: Score | None,
    ) -> Fixture:
        status = MatchStatus.FINISHED if score is not None else MatchStatus.SCHEDULED
        fixture = Fixture(
            competition_id=competition_id,
            home_team_id=home_id,
            away_team_id=away_id,
            kickoff=kickoff,
            status=status,
            score=score,
            external_id=ext,
            external_source=SOURCE,
        )
        existing = await self._fixtures.get_by_external_id(SOURCE, ext)
        if existing is not None:
            fixture.id = existing.id
            await self._fixtures.update(fixture)
            self._counters.skipped_duplicates += 1  # 已存在 → 不算新导入
        else:
            await self._fixtures.add(fixture)
            self._counters.matches_imported += 1
        if score is not None:
            self._counters.scores_imported += 1
        return fixture

    async def _import_odds(
        self, fixture_id: UUID, prices: dict[str, Decimal | None], captured_at: datetime
    ) -> None:
        bet365 = await self._bet365_bookmaker()
        for code in ("home", "draw", "away"):
            price = prices.get(code)
            if price is None or price <= Decimal("1"):
                continue
            snapshot = OddsSnapshot(
                fixture_id=fixture_id,
                bookmaker_id=bet365.id,
                selection=Selection(market=MarketType.MATCH_RESULT, code=code),
                odds=Odds(price),
                captured_at=captured_at,
            )
            if await self._odds.add_if_absent(snapshot):
                self._counters.odds_imported += 1
            else:
                self._counters.skipped_duplicates += 1

    async def import_row(self, col: dict[str, int], row: tuple[Any, ...], sheet: str) -> None:
        def cell(field: str) -> Any:
            idx = col.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        home = cell("home")
        away = cell("away")
        kickoff = parse_date(cell("date"))
        if not home or not away or kickoff is None or str(home) == str(away):
            self._counters.rows_invalid += 1
            return

        comp_name = str(cell("competition") or sheet).strip()
        season = cell("season")
        try:
            competition = await self._get_or_create_competition(comp_name, None)
            home_team = await self._upsert_team(str(home).strip(), None)
            away_team = await self._upsert_team(str(away).strip(), None)

            hs, as_ = to_int(cell("home_score")), to_int(cell("away_score"))
            score = Score(home=hs, away=as_) if hs is not None and as_ is not None else None

            ext_id = cell("external_id")
            ext = (
                str(ext_id).strip()
                if ext_id
                else deterministic_id(comp_name, season, kickoff.date().isoformat(), home, away)
            )
            fixture = await self._upsert_fixture(
                ext, competition.id, home_team.id, away_team.id, kickoff, score
            )

            await self._import_odds(
                fixture.id,
                {
                    "home": to_decimal(cell("b365_home")),
                    "draw": to_decimal(cell("b365_draw")),
                    "away": to_decimal(cell("b365_away")),
                },
                captured_at=kickoff,
            )

            # xG / shots / SoT：读取但不落库（无对应仓储）。
            if any(
                cell(f) not in (None, "")
                for f in ("xg_home", "xg_away", "shots_home", "shots_away", "sot_home", "sot_away")
            ):
                self._counters.stats_rows_read += 1
        except (ValueError, InvalidOperation) as exc:
            logger.warning("Skipping invalid row in %s (%s vs %s): %s", sheet, home, away, exc)
            self._counters.rows_invalid += 1

    async def import_workbook(self, path: Path) -> Counters:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = ws.iter_rows(values_only=True)
                header = next((r for r in rows if r and sum(c is not None for c in r) >= 2), None)
                if header is None:
                    continue
                col = build_column_map(header)
                if "home" not in col or "away" not in col or "date" not in col:
                    logger.warning("Sheet '%s' missing home/away/date columns; skipped", sheet)
                    continue
                for row in rows:
                    if row is None or all(c is None for c in row):
                        continue
                    await self.import_row(col, row, sheet)
        finally:
            wb.close()
        return self._counters


async def run(path: Path) -> Counters:
    db = Database(get_settings().sqlalchemy_dsn)
    try:
        async with db.session() as session:
            importer = WorldCupExcelImporter(session)
            return await importer.import_workbook(path)
    finally:
        await db.dispose()


def main() -> None:
    parser = argparse.ArgumentParser(description="Import a World Cup Excel workbook.")
    parser.add_argument("path", type=Path, help="Path to WorldCup2026.xlsx")
    args = parser.parse_args()
    if not args.path.exists():
        raise SystemExit(f"file not found: {args.path}")

    c = asyncio.run(run(args.path))
    print(f"Matches imported:   {c.matches_imported}")
    print(f"Scores imported:    {c.scores_imported}")
    print(f"Teams created:      {c.teams_created}")
    print(f"Teams updated:      {c.teams_updated}")
    print(f"Odds imported:      {c.odds_imported}")
    print(f"Skipped duplicates: {c.skipped_duplicates}")
    if c.rows_invalid:
        print(f"(invalid/empty rows skipped: {c.rows_invalid})")
    if c.stats_rows_read:
        print(
            f"(xG/shots/SoT rows read but NOT persisted: {c.stats_rows_read} "
            f"— no match-stats repository exists)"
        )


if __name__ == "__main__":
    main()
